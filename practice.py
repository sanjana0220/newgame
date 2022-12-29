from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
wb=load_workbook('iterate.xlsx')
ws= wb.active
for row in range(1,2):
  for col in range (1,2):
    char = get_column_letter(col)
    print(ws[char+str(row)].value)

